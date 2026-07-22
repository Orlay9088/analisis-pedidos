import openpyxl
import pandas as pd
from collections import defaultdict
from datetime import datetime
from typing import Optional


COLUMN_KEYWORDS = {
    "asesor": ["asesor", "vendedor", "nombre vendedor"],
    "canal_dist": ["canal distribucion", "canal dist"],
    "estado_mov": ["estado movto", "estado mov", "estado movto."],
    "fecha": ["fecha"],
    "cant_pedida": ["cant. pedida", "cant pedida", "cantidad pedida"],
    "cant_pendiente": ["cant. pendiente", "cant pendiente", "cantidad pendiente"],
    "cant_comprometida": ["cant. comprom", "cant comprom", "cantidad comprometida"],
    "linea": ["linea", "línea"],
    "sub_linea": ["sub-linea", "sub linea", "sublínea"],
    "cliente": ["cliente despacho", "cliente", "razon social"],
    "documento": ["nro documento", "nro doc", "documento"],
    "referencia": ["referencia"],
    "desc_item": ["desc. item", "desc item", "descripcion item", "item resumen"],
    "bodega": ["bodega"],
    "co": ["c.o.", "co"],
    "orden_compra": ["orden de compra", "orden compra"],
    "desc_sucursal": ["desc. sucursal despacho", "sucursal", "desc. sucursal"],
    "valor_pendiente_subtotal": ["valor pendiente subtotal", "valor subtotal"],
    "v_unidad": ["v.unidad", "v unidad"],
    "v_comprometido": ["v.comprometido", "v comprometido"],
    "valor_pendiente_neto": ["valor pendiente neto"],
    "valor_subtotal_local": ["valor subtotal local"],
    "lista_precios": ["lista de precios", "lista precios"],
    "tipo_cliente": ["tipo de cliente", "tipo cliente"],
    "utilidad": ["utilidad promedio", "utilidad"],
    "margen": ["margen promedio", "margen"],
    "desc_co": ["desc. c.o.", "desc c.o.", "desc. co"],
    "valor_descuentos": ["valor descuentos", "descuento"],
}


def _normalize(text: str) -> str:
    result = text.strip().lower().replace(".", " ").replace("_", " ")
    while "  " in result:
        result = result.replace("  ", " ")
    return result


def _match_column(header: str, keywords: list[str]) -> bool:
    norm = _normalize(header)
    return any(_normalize(kw) in norm for kw in keywords)


def detect_columns(headers: list[str]) -> dict[str, Optional[int]]:
    detected = {}
    for key, keywords in COLUMN_KEYWORDS.items():
        detected[key] = None
        for idx, h in enumerate(headers):
            if h and _match_column(str(h), keywords):
                detected[key] = idx
                break
    return detected


def find_data_sheet(wb: openpyxl.Workbook) -> tuple[Optional[str], Optional[str]]:
    pivot_sheet = None
    data_sheet = None
    max_rows = 0

    for name in wb.sheetnames:
        ws = wb[name]
        if ws.max_row and ws.max_row > max_rows:
            max_rows = ws.max_row
            data_sheet = name

    for name in wb.sheetnames:
        ws = wb[name]
        if ws.max_row and ws.max_row < 50 and ws.max_column and ws.max_column <= 15:
            pivot_sheet = name
            break

    return data_sheet, pivot_sheet


def read_pivot_table(ws) -> dict:
    rows = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=True):
        rows.append(list(row))

    title = None
    filter_value = None
    headers_row = None
    data_rows = []

    for i, row in enumerate(rows):
        non_none = [c for c in row if c is not None]
        if len(non_none) >= 2 and isinstance(non_none[0], str) and "canal" in _normalize(str(non_none[0])):
            title = str(non_none[0])
            filter_value = str(non_none[1]) if len(non_none) > 1 else None
            continue
        if any(c and isinstance(c, str) and "etiqueta" in _normalize(str(c)) for c in row if c):
            headers_row = row
            continue
        if any(c and isinstance(c, str) and "total" in _normalize(str(c)) for c in row if c):
            continue
        if headers_row is not None and any(c is not None for c in row):
            data_rows.append(row)

    return {
        "title": title,
        "filter_value": filter_value,
        "headers": headers_row,
        "data": data_rows,
    }


def _parse_pivot_from_df(df: pd.DataFrame) -> dict:
    title = None
    filter_value = None
    headers_row = None
    data_rows = []

    for i, row in df.iterrows():
        vals = [row[j] for j in range(len(df.columns))]
        non_none = [c for c in vals if pd.notna(c)]
        if len(non_none) >= 2 and isinstance(non_none[0], str) and "canal" in _normalize(str(non_none[0])):
            title = str(non_none[0])
            filter_value = str(non_none[1]) if len(non_none) > 1 else None
            continue
        if any(pd.notna(c) and isinstance(c, str) and "etiqueta" in _normalize(str(c)) for c in vals):
            headers_row = [c if pd.notna(c) else None for c in vals]
            continue
        if any(pd.notna(c) and isinstance(c, str) and "total" in _normalize(str(c)) for c in vals):
            continue
        if headers_row is not None and any(pd.notna(c) for c in vals):
            data_rows.append([c if pd.notna(c) else None for c in vals])

    return {
        "title": title,
        "filter_value": filter_value,
        "headers": headers_row,
        "data": data_rows,
    }


def _extract_pivot_filters(filepath: str) -> list[dict]:
    """Extrae filtros reales del cache del pivot table en Excel parseando XML directamente."""
    import zipfile
    import xml.etree.ElementTree as ET

    filters = []
    ns = {'x': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}

    try:
        with zipfile.ZipFile(filepath) as z:
            cache_files = [n for n in z.namelist()
                          if 'pivotcachedefinition' in n.lower() and n.endswith('.xml')]
            pt_files = [n for n in z.namelist()
                       if 'pivottable' in n.lower() and 'cache' not in n.lower() and n.endswith('.xml')]

            field_names = []
            shared_items_map = {}
            for cf in cache_files:
                cf_root = ET.fromstring(z.read(cf))
                for i, cf_elem in enumerate(cf_root.findall('.//x:cacheField', ns)):
                    fname = cf_elem.get('name', f'Field_{i}')
                    field_names.append(fname)
                    si = cf_elem.find('x:sharedItems', ns)
                    if si is not None:
                        items = []
                        for child in si:
                            v = child.get('v')
                            if v is not None:
                                items.append(v)
                        shared_items_map[i] = items

            for pt in pt_files:
                pt_root = ET.fromstring(z.read(pt))
                for pfield in pt_root.findall('.//x:pivotField', ns):
                    if pfield.get('axis') != 'axisPage':
                        continue
                    items_elem = pfield.find('x:items', ns)
                    if items_elem is None:
                        continue

                    pf_elems = pt_root.findall('.//x:pageFields/x:pageField', ns)
                    fname = 'Unknown'
                    fld_idx = -1
                    for pf_elem in pf_elems:
                        fld_idx = int(pf_elem.get('fld', -1))
                        if 0 <= fld_idx < len(field_names):
                            fname = field_names[fld_idx]
                            break

                    field_shared = shared_items_map.get(fld_idx, [])

                    selected = []
                    hidden_vals = []
                    all_values = []
                    for item in items_elem.findall('x:item', ns):
                        x = item.get('x')
                        h = item.get('h')
                        t = item.get('t')
                        if t == 'default' or x is None:
                            continue
                        x_idx = int(x)
                        val = field_shared[x_idx] if x_idx < len(field_shared) else str(x_idx)
                        all_values.append(val)
                        if h == '1':
                            hidden_vals.append(val)
                        else:
                            selected.append(val)

                    if selected or hidden_vals:
                        filters.append({
                            'field_name': fname,
                            'selected': selected,
                            'hidden': hidden_vals,
                            'all_values': all_values,
                        })

    except Exception:
        return []

    return filters


def read_raw_data(ws, col_map: dict[str, Optional[int]], filepath: str = None, sheet_name: str = None) -> pd.DataFrame:
    active_cols = {k: v for k, v in col_map.items() if v is not None}
    if not active_cols:
        return pd.DataFrame()

    if filepath and sheet_name:
        try:
            df = pd.read_excel(filepath, sheet_name=sheet_name, header=0, engine='openpyxl')
            rename_map = {}
            for key, idx in active_cols.items():
                if idx < len(df.columns):
                    rename_map[df.columns[idx]] = key
            df = df.rename(columns=rename_map)
            keep = [v for v in active_cols.values() if v < len(df.columns)]
            df = df.iloc[:, keep]
            df.columns = [k for k, v in active_cols.items() if v < len(df.columns)]
            return df
        except Exception:
            pass

    records = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        record = {}
        for key, idx in col_map.items():
            if idx is not None and idx < len(row):
                record[key] = row[idx]
            else:
                record[key] = None
        records.append(record)
    return pd.DataFrame(records)


def compute_asesor_metrics(df: pd.DataFrame, col_map: dict[str, Optional[int]]) -> list[dict]:
    if "asesor" not in df.columns or df["asesor"].isna().all():
        return []

    df_valid = df.dropna(subset=["asesor"]).copy()

    for col in ["asesor", "desc_sucursal", "documento", "desc_item", "desc_co", "linea", "sub_linea"]:
        if col in df_valid.columns:
            df_valid[col] = df_valid[col].astype(str).str.strip()
            df_valid[col] = df_valid[col].replace({"": None, "nan": None, "None": None})

    results = []
    for asesor_name, group in df_valid.groupby("asesor"):
        asesor_name = str(asesor_name).strip()
        if not asesor_name or asesor_name == "None":
            continue

        cant_pedida = _safe_sum(group, "cant_pedida")
        cant_pendiente = _safe_sum(group, "cant_pendiente")
        cant_comprometida = _safe_sum(group, "cant_comprometida")
        backlog_pct = (cant_pendiente / cant_pedida * 100) if cant_pedida > 0 else 0

        total_registros = len(group)

        valor_total = _safe_sum(group, "v_unidad")
        descuento_total = _safe_sum(group, "valor_descuentos")
        utilidad_promedio = _safe_mean(group, "utilidad")
        margen_promedio = _safe_mean(group, "margen")

        pedidos_por_estado = {}
        if "estado_mov" in group.columns:
            for estado, sub in group.groupby("estado_mov"):
                estado_str = str(estado).strip() if estado else "Sin estado"
                pedidos_por_estado[estado_str] = {
                    "registros": len(sub),
                    "cant_pedida": round(_safe_sum(sub, "cant_pedida"), 0),
                    "cant_pendiente": round(_safe_sum(sub, "cant_pendiente"), 0),
                    "cant_comprometida": round(_safe_sum(sub, "cant_comprometida"), 0),
                }

        top_lineas = []
        if "linea" in group.columns:
            linea_counts = group.dropna(subset=["linea"]).groupby("linea").agg(
                pedida=("cant_pedida", lambda x: _safe_sum_from_series(x)),
                registros=("linea", "count"),
            ).sort_values("pedida", ascending=False).head(5)
            for linea_name, row_data in linea_counts.iterrows():
                top_lineas.append({
                    "linea": str(linea_name),
                    "cant_pedida": round(row_data["pedida"], 0),
                    "registros": int(row_data["registros"]),
                })

        top_sub_lineas = []
        if "sub_linea" in group.columns:
            sub_counts = group.dropna(subset=["sub_linea"]).groupby("sub_linea").agg(
                pedida=("cant_pedida", lambda x: _safe_sum_from_series(x)),
                registros=("sub_linea", "count"),
            ).sort_values("pedida", ascending=False).head(5)
            for sub_name, row_data in sub_counts.iterrows():
                top_sub_lineas.append({
                    "sub_linea": str(sub_name),
                    "cant_pedida": round(row_data["pedida"], 0),
                    "registros": int(row_data["registros"]),
                })

        top_clientes = []
        if "cliente" in group.columns or "desc_sucursal" in group.columns:
            cliente_col = "desc_sucursal" if "desc_sucursal" in group.columns else "cliente"
            if cliente_col in group.columns:
                cliente_counts = group.dropna(subset=[cliente_col]).groupby(cliente_col).agg(
                    pedida=("cant_pedida", lambda x: _safe_sum_from_series(x)),
                    registros=(cliente_col, "count"),
                ).sort_values("pedida", ascending=False).head(5)
                for cliente_name, row_data in cliente_counts.iterrows():
                    top_clientes.append({
                        "cliente": str(cliente_name),
                        "cant_pedida": round(row_data["pedida"], 0),
                        "registros": int(row_data["registros"]),
                    })

        documentos_unicos = 0
        if "documento" in group.columns:
            documentos_unicos = group["documento"].dropna().nunique()

        pedidos_retenidos = pedidos_por_estado.get("Retenido", {}).get("registros", 0)
        pedidos_aprobados = pedidos_por_estado.get("Aprobado", {}).get("registros", 0)

        unidades_negocio = {}
        if "desc_co" in group.columns:
            co_counts = group.dropna(subset=["desc_co"]).groupby("desc_co").agg(
                pedida=("cant_pedida", lambda x: _safe_sum_from_series(x)),
                valor=("v_unidad", lambda x: _safe_sum_from_series(x)),
                registros=("desc_co", "count"),
            ).sort_values("pedida", ascending=False)
            for co_name, row_data in co_counts.iterrows():
                unidades_negocio[str(co_name)] = {
                    "cant_pedida": round(row_data["pedida"], 0),
                    "valor": round(row_data["valor"], 0),
                    "registros": int(row_data["registros"]),
                }

        top_estados_produccion = {}
        estado_col = None
        for col in group.columns:
            if _normalize(str(col)).strip() == "estado":
                estado_col = col
                break
        if estado_col:
            for estado, sub in group.groupby(estado_col):
                estado_str = str(estado).strip() if pd.notna(estado) else "Sin estado"
                top_estados_produccion[estado_str] = len(sub)

        top_proyectos = []
        documentos_por_proyecto = {}
        if "desc_sucursal" in group.columns:
            proj_df = group.dropna(subset=["desc_sucursal"])
            if not proj_df.empty:
                proj_agg = proj_df.groupby("desc_sucursal").agg(
                    pedida=("cant_pedida", lambda x: _safe_sum_from_series(x)),
                    pendiente=("cant_pendiente", lambda x: _safe_sum_from_series(x)),
                    comprometida=("cant_comprometida", lambda x: _safe_sum_from_series(x)),
                    valor_pend=("valor_pendiente_subtotal", lambda x: _safe_sum_from_series(x)) if "valor_pendiente_subtotal" in proj_df.columns else ("cant_pedida", lambda x: 0),
                    v_comprom=("v_comprometido", lambda x: _safe_sum_from_series(x)) if "v_comprometido" in proj_df.columns else ("cant_pedida", lambda x: 0),
                ).sort_values("pedida", ascending=False)
                for proj_name, row_data in proj_agg.iterrows():
                    proj_str = str(proj_name)
                    top_proyectos.append({
                        "proyecto": proj_str,
                        "cant_pedida": round(row_data["pedida"], 0),
                        "cant_pendiente": round(row_data["pendiente"], 0),
                        "cant_comprometida": round(row_data["comprometida"], 0),
                        "valor_pendiente": round(row_data["valor_pend"], 0),
                        "v_comprometido": round(row_data["v_comprom"], 0),
                    })
                    proj_docs_df = proj_df[proj_df["desc_sucursal"] == proj_name]
                    if "documento" in proj_docs_df.columns:
                        doc_groups = proj_docs_df.dropna(subset=["documento"]).groupby("documento")
                        docs_list = []
                        for doc_name, doc_sub in doc_groups:
                            doc_str = str(doc_name).strip()
                            items_list = []
                            if "desc_item" in doc_sub.columns:
                                items_df = doc_sub.dropna(subset=["desc_item"])
                                if not items_df.empty:
                                    items_agg = items_df.groupby("desc_item").agg(
                                        pedida=("cant_pedida", lambda x: _safe_sum_from_series(x)),
                                        pendiente=("cant_pendiente", lambda x: _safe_sum_from_series(x)),
                                        comprometida=("cant_comprometida", lambda x: _safe_sum_from_series(x)),
                                        valor_pend=("valor_pendiente_subtotal", lambda x: _safe_sum_from_series(x)) if "valor_pendiente_subtotal" in items_df.columns else ("cant_pedida", lambda x: 0),
                                        v_comprom=("v_comprometido", lambda x: _safe_sum_from_series(x)) if "v_comprometido" in items_df.columns else ("cant_pedida", lambda x: 0),
                                    ).sort_values("pedida", ascending=False)
                                    for item_name, irow in items_agg.iterrows():
                                        items_list.append({
                                            "item": str(item_name),
                                            "cant_pedida": round(irow["pedida"], 0),
                                            "cant_pendiente": round(irow["pendiente"], 0),
                                            "cant_comprometida": round(irow["comprometida"], 0),
                                            "valor_pendiente": round(irow["valor_pend"], 0),
                                            "v_comprometido": round(irow["v_comprom"], 0),
                                        })
                            docs_list.append({
                                "documento": doc_str,
                                "cant_pedida": round(_safe_sum(doc_sub, "cant_pedida"), 0),
                                "cant_pendiente": round(_safe_sum(doc_sub, "cant_pendiente"), 0),
                                "cant_comprometida": round(_safe_sum(doc_sub, "cant_comprometida"), 0),
                                "valor_pendiente": round(_safe_sum(doc_sub, "valor_pendiente_subtotal"), 0),
                                "v_comprometido": round(_safe_sum(doc_sub, "v_comprometido"), 0),
                                "items": items_list,
                            })
                        documentos_por_proyecto[proj_str] = docs_list

        desglose_contrato = {}
        if "documento" in group.columns:
            doc_df = group.dropna(subset=["documento"])
            for doc, sub in doc_df.groupby("documento"):
                doc_str = str(doc).strip()
                if doc_str.startswith("PDM"):
                    tipo = "Suministro"
                elif doc_str.startswith("PD-"):
                    tipo = "Instalacion"
                else:
                    tipo = "Otro"
                if tipo not in desglose_contrato:
                    desglose_contrato[tipo] = {"pedida": 0, "pendiente": 0, "comprometida": 0, "registros": 0}
                desglose_contrato[tipo]["pedida"] += round(_safe_sum(sub, "cant_pedida"), 0)
                desglose_contrato[tipo]["pendiente"] += round(_safe_sum(sub, "cant_pendiente"), 0)
                desglose_contrato[tipo]["comprometida"] += round(_safe_sum(sub, "cant_comprometida"), 0)
                desglose_contrato[tipo]["registros"] += len(sub)

        results.append({
            "asesor": asesor_name,
            "cant_pedida": round(cant_pedida, 0),
            "cant_pendiente": round(cant_pendiente, 0),
            "cant_comprometida": round(cant_comprometida, 0),
            "valor_pendiente": round(_safe_sum(group, "valor_pendiente_subtotal"), 0) if "valor_pendiente_subtotal" in group.columns else 0,
            "v_comprometido": round(_safe_sum(group, "v_comprometido"), 0) if "v_comprometido" in group.columns else 0,
            "backlog_pct": round(backlog_pct, 0),
            "total_registros": total_registros,
            "documentos_unicos": documentos_unicos,
            "pedidos_por_estado": pedidos_por_estado,
            "pedidos_retenidos": pedidos_retenidos,
            "pedidos_aprobados": pedidos_aprobados,
            "top_lineas": top_lineas,
            "top_sub_lineas": top_sub_lineas,
            "top_clientes": top_clientes,
            "valor_total": round(valor_total, 0),
            "utilidad_promedio": round(utilidad_promedio, 0),
            "margen_promedio": round(margen_promedio, 0),
            "descuento_total": round(descuento_total, 0),
            "unidades_negocio": unidades_negocio,
            "top_estados_produccion": top_estados_produccion,
            "top_proyectos": top_proyectos,
            "documentos_por_proyecto": documentos_por_proyecto,
            "desglose_contrato": desglose_contrato,
        })

    results.sort(key=lambda x: x["cant_pedida"], reverse=True)
    return results


def _rebuild_pivot_from_metrics(metrics: list[dict], canal_filter: str) -> dict:
    headers = ["Etiquetas de fila", "Suma de Cant. pedida", "Suma de Cant. pendiente",
               "Suma de Cant. comprom.", "Suma de Valor pendiente subtotal", "Suma de V.COMPROMETIDO"]
    data = []
    for m in metrics:
        data.append([
            m["asesor"], m["cant_pedida"], m["cant_pendiente"],
            m["cant_comprometida"], m["valor_pendiente"], m["v_comprometido"],
        ])
    total = ["Total",
             sum(m["cant_pedida"] for m in metrics),
             sum(m["cant_pendiente"] for m in metrics),
             sum(m["cant_comprometida"] for m in metrics),
             sum(m["valor_pendiente"] for m in metrics),
             sum(m["v_comprometido"] for m in metrics)]
    data.append(total)
    return {"title": "CANAL DISTRIBUCION", "filter_value": canal_filter, "headers": headers, "data": data}


def build_team_summary(metrics: list[dict]) -> dict:
    total_pedida = sum(m["cant_pedida"] for m in metrics)
    total_pendiente = sum(m["cant_pendiente"] for m in metrics)
    total_comprometida = sum(m["cant_comprometida"] for m in metrics)
    backlog_general = (total_pendiente / total_pedida * 100) if total_pedida > 0 else 0
    total_docs = sum(m["documentos_unicos"] for m in metrics)
    total_retenidos = sum(m["pedidos_retenidos"] for m in metrics)

    promedio_pedida = total_pedida / len(metrics) if metrics else 0
    promedio_pendiente = total_pendiente / len(metrics) if metrics else 0

    total_valor = sum(m["valor_total"] for m in metrics)
    total_descuentos = sum(m["descuento_total"] for m in metrics)
    margenes = [m["margen_promedio"] for m in metrics if m["margen_promedio"] > 0]
    utilidades = [m["utilidad_promedio"] for m in metrics if m["utilidad_promedio"] > 0]
    margen_promedio_equipo = sum(margenes) / len(margenes) if margenes else 0
    utilidad_promedio_equipo = sum(utilidades) / len(utilidades) if utilidades else 0

    unidades_negocio_total = {}
    for m in metrics:
        for co, data in m.get("unidades_negocio", {}).items():
            if co not in unidades_negocio_total:
                unidades_negocio_total[co] = {"cant_pedida": 0, "valor": 0, "registros": 0}
            unidades_negocio_total[co]["cant_pedida"] += data["cant_pedida"]
            unidades_negocio_total[co]["valor"] += data["valor"]
            unidades_negocio_total[co]["registros"] += data["registros"]

    return {
        "total_asesores": len(metrics),
        "total_pedida": round(total_pedida, 0),
        "total_pendiente": round(total_pendiente, 0),
        "total_comprometida": round(total_comprometida, 0),
        "backlog_general_pct": round(backlog_general, 0),
        "total_documentos_unicos": total_docs,
        "total_retenidos": total_retenidos,
        "promedio_pedida_por_asesor": round(promedio_pedida, 0),
        "promedio_pendiente_por_asesor": round(promedio_pendiente, 0),
        "total_valor": round(total_valor, 0),
        "total_descuentos": round(total_descuentos, 0),
        "margen_promedio_equipo": round(margen_promedio_equipo, 0),
        "utilidad_promedio_equipo": round(utilidad_promedio_equipo, 0),
        "unidades_negocio_total": unidades_negocio_total,
    }


def build_asesor_prompt(asesor: dict, team_summary: dict, canal: str = "") -> str:
    canal_info = f"\n- Canal de distribución: {canal}" if canal else ""

    avg_pedida = team_summary["promedio_pedida_por_asesor"]
    ratio_vs_promedio = (asesor["cant_pedida"] / avg_pedida * 100) if avg_pedida > 0 else 0

    lineas_text = ""
    if asesor["top_lineas"]:
        lineas_text = "\n".join(
            f"    - {l['linea']}: {l['cant_pedida']} unidades ({l['registros']} pedidos)"
            for l in asesor["top_lineas"]
        )
    else:
        lineas_text = "    (Sin datos de líneas)"

    sub_lineas_text = ""
    if asesor["top_sub_lineas"]:
        sub_lineas_text = "\n".join(
            f"    - {s['sub_linea']}: {s['cant_pedida']} unidades ({s['registros']} pedidos)"
            for s in asesor["top_sub_lineas"]
        )
    else:
        sub_lineas_text = "    (Sin datos de sub-líneas)"

    clientes_text = ""
    if asesor["top_clientes"]:
        clientes_text = "\n".join(
            f"    - {c['cliente']}: {c['cant_pedida']} unidades ({c['registros']} pedidos)"
            for c in asesor["top_clientes"]
        )
    else:
        clientes_text = "    (Sin datos de clientes)"

    estados_text = ""
    if asesor["pedidos_por_estado"]:
        estados_text = "\n".join(
            f"    - {estado}: {d['registros']} registros, pedida {d['cant_pedida']:,.0f}, pendiente {d['cant_pendiente']:,.0f}, comprometida {d['cant_comprometida']:,.0f}"
            for estado, d in asesor["pedidos_por_estado"].items()
        )

    co_text = ""
    if asesor.get("unidades_negocio"):
        co_text = "\n".join(
            f"    - {co}: {d['cant_pedida']} unidades, ${d['valor']:,.0f} valor, {d['registros']} registros"
            for co, d in asesor["unidades_negocio"].items()
        )
    else:
        co_text = "    (Sin datos de unidades de negocio)"

    estados_prod_text = ""
    if asesor.get("top_estados_produccion"):
        estados_prod_text = "\n".join(
            f"    - {estado}: {cant}" for estado, cant in asesor["top_estados_produccion"].items()
        )
    else:
        estados_prod_text = "    (Sin datos)"

    proyectos_text = ""
    if asesor.get("top_proyectos"):
        docs_por_proy = asesor.get("documentos_por_proyecto", {})
        for p in asesor["top_proyectos"]:
            proyectos_text += f"    - {p['proyecto']}: pedida {p['cant_pedida']:,.0f}, pendiente {p['cant_pendiente']:,.0f}, comprometida {p['cant_comprometida']:,.0f}, valor pendiente {p['valor_pendiente']:,.0f}, v.comprometido {p['v_comprometido']:,.0f}\n"
            docs = docs_por_proy.get(p['proyecto'], [])
            for doc in docs:
                proyectos_text += f"        - {doc['documento']}: pedida {doc['cant_pedida']:,.0f}, pendiente {doc['cant_pendiente']:,.0f}, comprometida {doc['cant_comprometida']:,.0f}, valor {doc['valor_pendiente']:,.0f}, v.comprom {doc['v_comprometido']:,.0f}\n"
                for it in doc.get('items', []):
                    proyectos_text += f"            - {it['item']}: pedida {it['cant_pedida']:,.0f}, pendiente {it['cant_pendiente']:,.0f}, comprometida {it['cant_comprometida']:,.0f}\n"
    else:
        proyectos_text = "    (Sin datos de proyectos)"

    items_text = ""
    doc_items = []
    for docs in (asesor.get("documentos_por_proyecto") or {}).values():
        for doc in docs:
            for it in doc.get("items", []):
                doc_items.append(it)
    if doc_items:
        doc_items.sort(key=lambda x: x.get("cant_pedida", 0), reverse=True)
        items_text = "\n".join(
            f"    - {i['item']}: pedida {i['cant_pedida']:,.0f}, pendiente {i['cant_pendiente']:,.0f}, comprometida {i['cant_comprometida']:,.0f}"
            for i in doc_items
        )
    else:
        items_text = "    (Sin datos de items)"

    contrato_text = ""
    if asesor.get("desglose_contrato"):
        for tipo, d in asesor["desglose_contrato"].items():
            contrato_text += f"    - {tipo}: pedida {d['pedida']:,.0f}, pendiente {d['pendiente']:,.0f}, comprometida {d['comprometida']:,.0f} ({d['registros']} registros)\n"
        contrato_text = contrato_text.rstrip()
    else:
        contrato_text = "    (Sin datos de contratos)"

    team_co_text = ""
    if team_summary.get("unidades_negocio_total"):
        team_co_text = "\n".join(
            f"    - {co}: {d['cant_pedida']} unidades, ${d['valor']:,.0f} valor total"
            for co, d in team_summary["unidades_negocio_total"].items()
        )

    valor_total_equipo = team_summary.get('total_valor', 0)
    valor_ratio = (asesor['valor_total'] / valor_total_equipo * 100) if valor_total_equipo > 0 else 0

    prompt = f"""Eres un analista experto en ventas y gestión de pedidos para una empresa de muebles y construcción. Genera un informe detallado y profesional para el siguiente asesor.

## DATOS DEL ASESOR
- Nombre: {asesor['asesor']}{canal_info}

## MÉTRICAS PRINCIPALES (CANTIDADES)
- Cantidad pedida total: {asesor['cant_pedida']:,.0f}
- Cantidad pendiente: {asesor['cant_pendiente']:,.0f}
- Cantidad comprometida: {asesor['cant_comprometida']:,.0f}
- % de Backlog (pendiente/pedida): {asesor['backlog_pct']}%
- Total de registros en sistema: {asesor['total_registros']}
- Documentos/pedidos únicos: {asesor['documentos_unicos']}
- Pedidos retenidos: {asesor['pedidos_retenidos']}
- Pedidos aprobados: {asesor['pedidos_aprobados']}
- Rendimiento vs promedio del equipo: {ratio_vs_promedio:.0f}%

## VALORES MONETARIOS Y RENTABILIDAD
- Valor total de los pedidos (V.UNIDAD): ${asesor['valor_total']:,.0f}
- Porcentaje del valor total del equipo: {valor_ratio:.1f}%
- Utilidad promedio por pedido: ${asesor['utilidad_promedio']:,.0f}
- Margen promedio: {asesor['margen_promedio']:.0f}%
- Descuentos totales aplicados: ${asesor['descuento_total']:,.0f}

## COMPARACIÓN CON EL EQUIPO
- Promedio del equipo (cantidad pedida): {avg_pedida:,.0f}
- Promedio del equipo (cantidad pendiente): {team_summary['promedio_pendiente_por_asesor']:,.0f}
- Backlog general del equipo: {team_summary['backlog_general_pct']}%
- Total pedida del equipo: {team_summary['total_pedida']:,.0f}
- Valor total del equipo: ${team_summary.get('total_valor', 0):,.0f}
- Margen promedio del equipo: {team_summary.get('margen_promedio_equipo', 0):.0f}%
- Utilidad promedio del equipo: ${team_summary.get('utilidad_promedio_equipo', 0):,.0f}

## ESTADOS DE PEDIDOS
{estados_text}

## ESTADOS DE PRODUCCIÓN / ESTADO GENERAL
{estados_prod_text}

## UNIDADES DE NEGOCIO (Desc. C.O.)
{co_text}

## UNIDADES DE NEGOCIO - EQUIPO COMPLETO
{team_co_text}

## TOP 5 LÍNEAS DE PRODUCTO (por cantidad pedida)
{lineas_text}

## TOP 5 SUB-LÍNEAS DE PRODUCTO
{sub_lineas_text}

## TOP 5 CLIENTES PRINCIPALES
{clientes_text}

## TODOS LOS PROYECTOS (por cantidad pedida)
{proyectos_text}

## TODOS LOS ITEMS (por cantidad pedida)
{items_text}

## TIPO DE CONTRATO (Instalación vs Suministro)
{contrato_text}

---

## INSTRUCCIONES
Genera un informe profesional usando formato Markdown. IMPORTANTE: Usa tablas Markdown para presentar datos numéricos. Ejemplo:
| Métrica | Valor |
|---|---|
| Cantidad pedida | 3,255 |

### 1. RESUMEN EJECUTIVO
Tabla resumen con los datos principales del asesor en formato | Métrica | Valor |.

### 2. ANÁLISIS DE RENDIMIENTO
Tabla comparativa: Asesor vs Promedio Equipo. Columnas: Métrica, Asesor, Promedio Equipo, Diferencia.

### 3. ANÁLISIS DEL BACKLOG
Tabla de backlog por estado. Análisis textual breve.

### 4. VALORES MONETARIOS Y RENTABILIDAD
Tabla con valores: Categoría, Valor Asesor, Valor Equipo, % Participación.

### 5. ANÁLISIS DE PROYECTOS, DOCUMENTOS E ITEMS
Tabla de los 5-10 proyectos principales con sus 5 métricas. Mencionar documentos e items clave.

### 6. UNIDADES DE NEGOCIO Y PRODUCTOS
Tabla de unidades de negocio y tabla de top líneas/productos.

### 7. PEDIDOS RETENIDOS Y ALERTAS
Tabla de estados con problemas. Alertas en lista.

### 8. RECOMENDACIONES ESPECÍFICAS
Tabla con: #, Recomendación, Acción, Prioridad, Impacto Esperado.

Sé directo, profesional y basado en datos numéricos específicos."""

    return prompt


def _safe_sum(df: pd.DataFrame, col: str) -> float:
    if col not in df.columns:
        return 0.0
    return float(pd.to_numeric(df[col], errors="coerce").fillna(0).sum())


def _safe_mean(df: pd.DataFrame, col: str) -> float:
    if col not in df.columns:
        return 0.0
    val = pd.to_numeric(df[col], errors="coerce").fillna(0).mean()
    return float(val) if pd.notna(val) else 0.0


def _safe_sum_from_series(series: pd.Series) -> float:
    return float(pd.to_numeric(series, errors="coerce").fillna(0).sum())


def process_excel(filepath: str, canal_filter_override: str = None) -> dict:
    try:
        xls = pd.ExcelFile(filepath, engine='openpyxl')
    except Exception as e:
        raise ValueError(f"No se pudo abrir el archivo Excel. Verifica que no este corrupto o abierto en otro programa: {e}")

    sheet_names = xls.sheet_names
    if not sheet_names:
        raise ValueError("El archivo Excel no tiene hojas.")

    data_sheet_name = None
    pivot_sheet_name = None
    max_rows = 0

    sheet_dfs = {}
    for name in sheet_names:
        try:
            df_check = pd.read_excel(filepath, sheet_name=name, header=None, engine='openpyxl')
            sheet_dfs[name] = df_check
            row_count = len(df_check)
            if row_count > max_rows:
                max_rows = row_count
                data_sheet_name = name
        except Exception:
            continue

    for name, df_check in sheet_dfs.items():
        if name != data_sheet_name and 0 < len(df_check) < 50 and len(df_check.columns) <= 15:
            pivot_sheet_name = name
            break

    pivot_data = None
    pivot_filters = []
    if pivot_sheet_name:
        try:
            df_pivot = pd.read_excel(filepath, sheet_name=pivot_sheet_name, header=None, engine='openpyxl')
            pivot_data = _parse_pivot_from_df(df_pivot)
        except Exception:
            pivot_data = None
        try:
            pivot_filters = _extract_pivot_filters(filepath)
        except Exception:
            pivot_filters = []

    if not data_sheet_name:
        raise ValueError("No se encontro una hoja con datos. Verifica que el archivo tenga al menos una hoja con datos.")

    headers = []
    try:
        df_raw = pd.read_excel(filepath, sheet_name=data_sheet_name, header=0, engine='openpyxl')
        headers = list(df_raw.columns)
        headers = [str(h) if h else '' for h in headers]
    except Exception as e:
        raise ValueError(f"Error al leer la hoja '{data_sheet_name}': {e}")

    if not headers:
        raise ValueError("La hoja de datos no tiene encabezados en la primera fila.")

    col_map = detect_columns(headers)

    rename_map = {}
    for key, idx in col_map.items():
        if idx is not None and idx < len(df_raw.columns):
            rename_map[df_raw.columns[idx]] = key
    df_raw = df_raw.rename(columns=rename_map)

    canal_dist_values = []
    if "canal_dist" in df_raw.columns:
        canal_dist_values = sorted([v for v in df_raw["canal_dist"].dropna().unique().tolist() if str(v).strip()])

    canal_filter = ""
    if pivot_data and pivot_data.get("filter_value"):
        canal_filter = pivot_data["filter_value"]

    for pf in pivot_filters:
        if pf["field_name"].upper().replace(" ", "").replace(".", "") in ["CANALDISTRIBUCION", "CANALDIST"]:
            if pf["selected"]:
                canal_filter = pf["selected"][0]
            break

    if canal_filter_override:
        canal_filter = canal_filter_override
    elif not canal_filter and canal_dist_values:
        canal_filter = canal_dist_values[0]

    df_filtered = df_raw
    if canal_filter and "canal_dist" in df_raw.columns:
        df_filtered = df_raw[df_raw["canal_dist"].astype(str).str.strip() == canal_filter].copy()

    asesor_metrics = compute_asesor_metrics(df_filtered, col_map)
    team_summary = build_team_summary(asesor_metrics)

    pivot_table = _rebuild_pivot_from_metrics(asesor_metrics, canal_filter)

    return {
        "pivot_table": pivot_table,
        "pivot_filters": pivot_filters,
        "col_map": {k: v for k, v in col_map.items() if v is not None},
        "asesor_metrics": asesor_metrics,
        "team_summary": team_summary,
        "canal_dist_values": canal_dist_values,
        "canal_filter": canal_filter,
        "total_raw_rows": len(df_filtered),
        "total_unfiltered_rows": len(df_raw),
        "detected_columns": {k: (headers[v] if v is not None else None) for k, v in col_map.items()},
    }
